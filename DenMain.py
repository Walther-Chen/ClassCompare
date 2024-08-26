# -*- coding: utf-8 -*-
"""
Created on Mon Aug 26 10:03:18 2024

@author: Tristan
"""

import classutility as util # Import the utility module
import json # Import the json module
import classtask as task # Import the task module
import openpyxl
import os

def dendict (medthree, medfour):
    """
    medthree與medfour為醫三與醫四課表的檔案名稱
    """
    afile= openpyxl.load_workbook(medthree)
    asheet=afile.worksheets[0]
    astor=task.task(asheet)
    bfile=openpyxl.load_workbook(medfour)
    bsheet=bfile.worksheets[0]
    bstor=task.task(bsheet)
    astor=task.teachers(astor)
    bstor=task.teachers(bstor)    
    d=dict()
    Yidate=task.denDate()
    for y in Yidate:
        d[y]=[]
    for j in astor:
        if j[1]=='一)'and j[2]=='A':
            if j[3]=='8-10' or j[3]=='10-12' or j[3]=='10-11' or j[3]=='8-9' or j[3]=='9-10':
                d[j[0]].append((j[3],j[4],'四'))
    for k in bstor:
        if k[1]=='一)'and k[2]=='A':
            if k[3]=='8-10' or k[3]=='10-12' or k[3]=='10-11' or k[3]=='8-9' or k[3]=='9-10':
                d[j[0]].append((k[3],k[4],'三'))
    return d

def denlist_generation (medthree, medfour):
    """
    中乙空白課表產生模組。medthree與medfour為醫三與醫四課表的檔案名稱
    """
    r=os.path.dirname(__file__)
    os.chdir(r) #reduce the possibility of FileNotFoundError
    d = dendict(medthree, medfour)
    li = []
    li.append(["授課日期", "時間", "授課教師", "授課內容", "同日醫學系上課時間與教師", "備註"])
    for k, v in d.items():
        #combine all the classes in the same day (v) into a string described in the format of (時間,老師,年級)
        medclassess = []
        for item in v:
            if not item[1]:
                medclassess.append(f"醫{item[2]}, {item[0]}節, PBL")
            else:
                medclassess.append(f"醫{item[2]}, {item[0]}節, {item[1]}")
        li.append([k, "8-10", "", "", ";".join(medclassess), ""])
    return li


def teacher_menu (filename):
    """
        A two level CLI menu for teachers and their classes, using multiple choice selection to select teacher and subject
    """
    loaded_teachers = util.load_teachers(filename)
    while True:
        # Create a list of keys in the loaded_teachers dictionary first, and display them with index numbers
        teacher_list = list(loaded_teachers.keys())
        print("Teachers:")
        for index, teacher in enumerate(teacher_list, start=1):
            print(f"{index}. {teacher}")
        choice = input("請選擇你要排課的老師: ")
        #list the classes of the selected teacher, and display them with index numbers
        try:
            selected_teacher = teacher_list[int(choice) - 1]
        except IndexError:
            print("Invalid choice. Please try again.")
            continue
        print(f"Classes for {selected_teacher}:")
        for index, subject in enumerate(loaded_teachers[selected_teacher], start=1):
            print(f"{index}. {subject}")
        choice = input("請選擇你要排課的科目: ")
        try: 
            selected_subject = loaded_teachers[selected_teacher][int(choice) - 1]
        except IndexError:
            print("Invalid choice. Please try again.")
            continue
        # Display the selected teacher and subject
        print(f"授課教師: {selected_teacher}")
        print(f"授課內容: {selected_subject}")
        # Ask if the user wants to continue or exit
        choice = input("修改請按 0, 離開請按其他鍵: ")
        if choice.lower() != '0':
            return selected_subject, selected_teacher
        
def main_menu(classchart):
    print("牙醫排課系統")
    print("目前課表：")
    util.print_chungyi(classchart)
    choice = input("請用序號選擇你要排哪一堂課，輸入q離開並存檔")

    if choice == "q":
        print("Exiting the program...")
        # Save the classchart to a excel file, do not use content manager
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in classchart:
            ws.append(row)
        wb.save("d_classchart.xlsx")
        print("課表已經存到 d_classchart.xlsx")
        return

    elif choice.isdigit():
        choice = int(choice)
        if choice > 0 and choice < len(classchart):            
            classchart[choice][3], classchart[choice][2] = teacher_menu("teachers.json")            
            main_menu(classchart)
        else:
            print("Invalid choice. Please try again.")
            main_menu(classchart)


# Start the program
classchart = denlist_generation("a.xlsx", "b.xlsx")
main_menu(classchart)