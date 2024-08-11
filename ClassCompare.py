# -*- coding: utf-8 -*-
"""
Created on Tue Jul 30 19:10:21 2024

@author: Tristan
"""

import openpyxl
import os
import classtask as task


def yidict (medthree, medfour):
    """
    medthree與medfour為醫三與醫四課表的檔案名稱
    """
    afile= openpyxl.load_workbook(medthree)
    asheet=afile.worksheets[0]
    astor=task.task(asheet)
    bfile=openpyxl.load_workbook(medfour)
    bsheet=bfile.worksheets[0]
    bstor=task.task(bsheet)
    astor=task.teachers(astor)
    bstor=task.teachers(bstor)    
    d=dict()
    Yidate=task.ChineseYiDate()
    for y in Yidate:
        d[y]=[]
    for j in astor:
        if j[1]=='一)'and j[2]=='P':
            if j[3]=='3-4' or j[3]=='4-5' or j[3]=='3-5':
                d[j[0]].append((j[3],j[4],'四'))
    for k in bstor:
        if k[1]=='一)'and k[2]=='P':
            if k[3]=='3-4' or k[3]=='4-5' or k[3]=='3-5':
                d[j[0]].append((k[3],k[4],'三'))
    return d

def yilist_generation (medthree, medfour):
    """
    中乙空白課表產生模組。medthree與medfour為醫三與醫四課表的檔案名稱
    """
    d = yidict(medthree, medfour)
    li = []
    li.append(["授課日期", "時間", "授課教師", "授課內容", "同日醫學系上課時間與教師", "備註"])
    for k, v in d.items():
        #combine all the classes in the same day (v) into a string described in the format of (時間,老師,年級)
        medclassess = []
        for item in v:
            if not item[1]:
                medclassess.append(f"醫{item[2]}, {item[0]}節, PBL")
            else:
                medclassess.append(f"醫{item[2]}, {item[0]}節, {item[1]}")
        li.append([k, "3-5", "", "", ";".join(medclassess), ""])
    return li
            
#data= r'C:\\Users\\Tristan\\OneDrive\\桌面\\code\\Python\\ClassCompare'
#os.chdir(data)
afile= openpyxl.load_workbook('a.xlsx')
asheet=afile.worksheets[0]
astor=task.task(asheet)

#b

bfile=openpyxl.load_workbook('b.xlsx')
bsheet=bfile.worksheets[0]
bstor=task.task(bsheet)

#攤開老師
astor=task.teachers(astor)
bstor=task.teachers(bstor)



#檢查是否撞課
for i in astor:
    if i in bstor:
        if None not in i:
            print(i)
            
#ChineseYi
wb = openpyxl.Workbook()
wb.create_sheet("中乙",0) 
s1= wb['中乙']

#處理所有在週一的課
d=dict()
Yidate=task.ChineseYiDate()
for y in Yidate:
    d[y]=[]

for j in astor:
    if j[1]=='一)'and j[2]=='P':
        if j[3]=='3-4' or j[3]=='4-5' or j[3]=='3-5':
            d[j[0]].append((j[3],j[4],'四'))
for k in bstor:
    if k[1]=='一)'and k[2]=='P':
        if k[3]=='3-4' or k[3]=='4-5' or k[3]=='3-5':
            d[j[0]].append((k[3],k[4],'三'))
print(d)
#匯出結果
s1.cell(1,1).value='日期'
s1.cell(1,2).value='(時間,老師,年級)'
ini=2
for write_class in d.items():
    s1.cell(ini,1).value=write_class[0]
    s1.cell(ini,2).value=str(write_class[1])
    ini=ini+1
wb.save('chineseyi.xlsx')







