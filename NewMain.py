# -*- coding: utf-8 -*-
"""
Created on Mon Aug 26 10:03:18 2024

@author: Tristan
"""
chinese_weekday = ['一)', '二)', '三)', '四)', '五)', '六)', '日)']
import classutility as util # Import the utility module
import json # Import the json module
import classtask as task # Import the task module
import openpyxl
import os
import datetime
from sys import exit

def datestring_to_datetime(datestring):
    """
    Transform datestring "113/9/11" to datetime(2024, 9, 11)
    """
    date_list = datestring.split('/')
    year = int(date_list[0]) + 1911
    month = int(date_list[1])
    day = int(date_list[2])
    return datetime.date(year, month, day)

def class_dict (medthree, medfour, weekday, start_time):
    """
    medthree與medfour為醫三與醫四課表的檔案名稱, weekday為上課的星期幾 (0 = 週一, 1 = 週二, ..., 6 = 週日), start_time為上課開始的時間
    """
    afile= openpyxl.load_workbook(medthree)
    asheet=afile.worksheets[0]
    astor=task.task(asheet)
    bfile=openpyxl.load_workbook(medfour)
    bsheet=bfile.worksheets[0]
    bstor=task.task(bsheet)
    astor=task.teachers(astor)
    bstor=task.teachers(bstor)    
    d=dict()
    academic_year, semester, start_date = task.semester_start_date()
    class_date = task.generate_course_dates_with_holidays(start_date, weekday, academic_year+1911)    
    if start_time <= 12:
        timing = "A"
    else:
        timing = "P"
    for y in class_date:
        d[y[0]]=[y[1]]
    for j in astor:
        if j[1]==chinese_weekday[weekday] and len(j) > 3 and j[2]==timing:            
            try:
                d[datestring_to_datetime(j[0])].append((j[3],j[4],'四'))
            except KeyError:
                pass
    for k in bstor:
        if k[1]==chinese_weekday[weekday] and len(j) > 3 and j[2]==timing:            
            try:
                d[datestring_to_datetime(j[0])].append((j[3],j[4],'三'))
            except KeyError:
                pass
    return d

def classlist_generation (medthree, medfour, weekday, start_time):
    """
    任意空白課表產生模組。medthree與medfour為醫三與醫四課表的檔案名稱
    """
    r=os.path.dirname(__file__)
    os.chdir(r) #reduce the possibility of FileNotFoundError
    d = class_dict(medthree, medfour, weekday, start_time)
    #print(d)
    #print(d.items())
    li = []
    li.append(["授課日期", "開始時間", "授課教師", "授課內容", "同日醫學系上課時間與教師", "國定假日"])
    for k, v in d.items():
        #combine all the classes in the same day (v) into a string described in the format of (時間,老師,年級)
        medclassess = []
        if len(v) > 1: #[None, ('1-3', '謝佩伶', '四'), ('1-3', '曾慶三', '四'), ('1-3', '吳政訓', '四'), ('1-3', '柯妙華', '四'), ('3-5', '謝佩伶', '四'), ('3-5', '曾慶三', '四'), ('3-5', '吳政訓', '四'), ('3-5', '柯妙華', '四')]
            #for item in v except the first item
            for item in v[1:]: #第一個項目是國定假日
                if not item[1]:
                    medclassess.append(f"醫{item[2]}, {item[0]}節, PBL")
                else:
                    medclassess.append(f"醫{item[2]}, {item[0]}節, {item[1]}")
        li.append([k, str(start_time) + ":00", "", "", ";".join(medclassess), v[0] if v[0] is not None else ""])
    return li


def teacher_menu (filename):
    """
        A two level CLI menu for teachers and their classes, using multiple choice selection to select teacher and subject
    """
    loaded_teachers = util.load_teachers(filename)
    while True:
        # Create a list of keys in the loaded_teachers dictionary first, and display them with index numbers
        teacher_list = list(loaded_teachers.keys())
        print("Teachers:")
        for index, teacher in enumerate(teacher_list, start=1):
            print(f"{index}. {teacher}")
        choice = input("請選擇你要排課的老師: ")
        #list the classes of the selected teacher, and display them with index numbers
        try:
            selected_teacher = teacher_list[int(choice) - 1]
        except IndexError:
            print("Invalid choice. Please try again.")
            continue
        print(f"Classes for {selected_teacher}:")
        for index, subject in enumerate(loaded_teachers[selected_teacher], start=1):
            print(f"{index}. {subject}")
        choice = input("請選擇你要排課的科目: ")
        try: 
            selected_subject = loaded_teachers[selected_teacher][int(choice) - 1]
        except IndexError:
            print("Invalid choice. Please try again.")
            continue
        # Display the selected teacher and subject
        print(f"授課教師: {selected_teacher}")
        print(f"授課內容: {selected_subject}")
        # Ask if the user wants to continue or exit
        choice = input("修改請按 0, 離開請按其他鍵: ")
        if choice.lower() != '0':
            return selected_subject, selected_teacher
        
def main_menu(classchart = None, weekday = None, start_time = None, classname = None):
    current_year, sementer, start_date = task.semester_start_date()
    print("中國醫藥大學排課系統，目前時間為", current_year, "學年度", sementer, "學期", "開學日期", start_date)
    if classchart is None:
        print("請選擇你要排課的系所：")
        # Display the list of departments read from classlist.json
        with open("courselist.json", 'r', encoding="utf-8") as json_file:
            course_list = json.load(json_file)
        for department in course_list:                                    
            print(f"{course_list.index(department)+1}. {department[util.CLASSNAME]} (星期{chinese_weekday[department[util.WEEKDAY]][0]} {department[util.STARTTIME]}:00開始)")
        choice = input("請用序號選擇你要排哪一個系所，輸入q離開: ")
        if choice == "q":
            print("Exiting the program...")
            exit(0)
        elif choice.isdigit():
            choice = int(choice)
            if choice > 0 and choice <= len(course_list):
                classchart = classlist_generation("a.xlsx", "b.xlsx", course_list[choice-1][util.WEEKDAY], course_list[choice-1][util.STARTTIME])
                main_menu(classchart, course_list[choice-1][util.WEEKDAY], course_list[choice-1][util.STARTTIME], course_list[choice-1][util.CLASSNAME])
            else:
                print("Invalid choice. Please try again.")
                main_menu()            

    util.print_chungyi(classchart)
    choice = input("請用序號選擇你要排哪一堂課，輸入q離開並存檔: ")
    outputfilename = f"{current_year}學年度{sementer}學期{classname}課表.xlsx"
    if choice == "q":
        print("Exiting the program...")
        # Save the classchart to a excel file, do not use content manager
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in classchart:
            ws.append(row)
        wb.save(outputfilename)
        print("課表已經存到", outputfilename)
        exit(0)

    elif choice.isdigit():
        choice = int(choice)
        if choice > 0 and choice < len(classchart):            
            classchart[choice][3], classchart[choice][2] = teacher_menu("teachers.json")            
            main_menu(classchart, weekday, start_time, classname)
        else:
            print("Invalid choice. Please try again.")
            main_menu(classchart, weekday, start_time, classname)


# Start the program
r=os.path.dirname(__file__)
os.chdir(r)
#print(class_dict("a.xlsx", "b.xlsx", 1, 13))
#classchart = classlist_generation("a.xlsx", "b.xlsx", 1, 13)
#print(classchart)
main_menu()
